"""
Router de Reportes — CCPL
Endpoints para Registro de Ventas SUNAT, Reporte de Caja y Resumen Mensual.
Incluye export a Excel.

Agregar a main.py:
    from app.routers.reportes import router as reportes_router
    app.include_router(reportes_router)

Agregar ruta del template:
    @app.get("/admin/reportes")
    async def admin_reportes(request: Request):
        return templates.TemplateResponse("pages/reportes.html", {"request": request})
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, and_, case
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Optional
from io import BytesIO

from app.database import get_db
from app.models import Payment, Organization, Colegiado

router = APIRouter(prefix="/api/reportes", tags=["reportes"])

PERU_TZ_OFFSET = timedelta(hours=-5)


def _rango_mes(mes: str, anio: str):
    """Retorna (inicio, fin) del mes dado."""
    m, a = int(mes), int(anio)
    inicio = datetime(a, m, 1)
    if m == 12:
        fin = datetime(a + 1, 1, 1)
    else:
        fin = datetime(a, m + 1, 1)
    return inicio, fin


def _rango_dia(fecha: str):
    """Retorna (inicio, fin) del día dado."""
    dia = datetime.strptime(fecha, "%Y-%m-%d")
    return dia.replace(hour=0, minute=0, second=0), dia + timedelta(days=1)


# ══════════════════════════════════════════════════════════
# REGISTRO DE VENTAS (Formato 14.1 SUNAT)
# ══════════════════════════════════════════════════════════

@router.get("/registro-ventas")
async def registro_ventas(
    mes: str,
    anio: str,
    centro_costo_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    inicio, fin = _rango_mes(mes, anio)

    query = db.query(Payment).filter(
        Payment.status.in_(["approved", "anulado"]),
        Payment.reviewed_at >= inicio,
        Payment.reviewed_at < fin,
    ).order_by(Payment.reviewed_at)

    pagos = query.all()

    registros = []
    for p in pagos:
        # Determinar tipo comprobante
        es_factura = (p.pagador_tipo == "empresa")
        tipo_comp = "01" if es_factura else "03"

        # Datos del cliente
        if es_factura:
            cli_tipo = "6"
            cli_doc = p.pagador_documento or ""
            cli_nombre = p.pagador_nombre or ""
        elif p.colegiado_id:
            col = db.query(Colegiado).filter(Colegiado.id == p.colegiado_id).first()
            cli_tipo = "1"
            cli_doc = col.dni if col else ""
            cli_nombre = col.apellidos_nombres if col else ""
        else:
            cli_tipo = "0"
            cli_doc = "-"
            cli_nombre = "PÚBLICO GENERAL"

        monto = float(p.amount or 0)

        # IGV: por defecto exonerado (CCPL tipo_afectacion = 20)
        base_imponible = 0.0
        exonerado = monto
        igv = 0.0

        registros.append({
            "fecha_emision": p.reviewed_at.strftime("%d/%m/%Y") if p.reviewed_at else "",
            "tipo_comprobante": tipo_comp,
            "serie": "F001" if es_factura else "B001",
            "numero": str(p.id).zfill(8),
            "cliente_tipo_doc": cli_tipo,
            "cliente_num_doc": cli_doc,
            "cliente_nombre": cli_nombre,
            "base_imponible": base_imponible,
            "exonerado": exonerado,
            "igv": igv,
            "total": monto,
            "estado": p.status,
            "metodo_pago": p.payment_method or "",
            "ref_nota_credito": "",
        })

    return {"registros": registros, "mes": mes, "anio": anio}


@router.get("/registro-ventas/excel")
async def registro_ventas_excel(
    mes: str,
    anio: str,
    centro_costo_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    data = await registro_ventas(mes, anio, centro_costo_id, db)
    registros = data["registros"]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Ventas"

    # ── Encabezado SUNAT ──
    ws.merge_cells("A1:O1")
    ws["A1"] = "REGISTRO DE VENTAS E INGRESOS"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:O2")
    ws["A2"] = "COLEGIO DE CONTADORES PÚBLICOS DE LORETO — RUC: 20103830991"
    ws["A2"].font = Font(size=11, name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")

    meses = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    ws.merge_cells("A3:O3")
    ws["A3"] = f"Período: {meses[int(mes)]} {anio}"
    ws["A3"].font = Font(size=11, name="Arial")
    ws["A3"].alignment = Alignment(horizontal="center")

    # ── Headers tabla ──
    headers = [
        "N°", "Fecha Emisión", "Tipo", "Serie", "Número",
        "Tipo Doc.", "Nro. Doc.", "Cliente",
        "Base Imponible", "Exonerado", "IGV", "Total",
        "Estado", "Método Pago", "Ref. NC"
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # ── Data rows ──
    money_fmt = '#,##0.00'
    for i, r in enumerate(registros, 1):
        row = i + 5
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=r["fecha_emision"]).border = thin_border
        ws.cell(row=row, column=3, value=r["tipo_comprobante"]).border = thin_border
        ws.cell(row=row, column=4, value=r["serie"]).border = thin_border
        ws.cell(row=row, column=5, value=r["numero"]).border = thin_border
        ws.cell(row=row, column=6, value=r["cliente_tipo_doc"]).border = thin_border
        ws.cell(row=row, column=7, value=r["cliente_num_doc"]).border = thin_border
        ws.cell(row=row, column=8, value=r["cliente_nombre"]).border = thin_border

        for c, val in [(9, r["base_imponible"]), (10, r["exonerado"]), (11, r["igv"]), (12, r["total"])]:
            cell = ws.cell(row=row, column=c, value=val)
            cell.number_format = money_fmt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

        ws.cell(row=row, column=13, value=r["estado"]).border = thin_border
        ws.cell(row=row, column=14, value=r["metodo_pago"]).border = thin_border
        ws.cell(row=row, column=15, value=r["ref_nota_credito"]).border = thin_border

        # Estilo anulado
        if r["estado"] == "anulado":
            for c in range(1, 16):
                ws.cell(row=row, column=c).font = Font(color="FF0000", strikethrough=True, name="Arial", size=10)

    # ── Fila totales ──
    if registros:
        tot_row = len(registros) + 6
        ws.cell(row=tot_row, column=8, value="TOTALES").font = Font(bold=True, name="Arial", size=10)
        activos = [r for r in registros if r["estado"] != "anulado"]
        for c, key in [(9, "base_imponible"), (10, "exonerado"), (11, "igv"), (12, "total")]:
            cell = ws.cell(row=tot_row, column=c)
            cell.value = sum(r[key] for r in activos)
            cell.number_format = money_fmt
            cell.font = Font(bold=True, name="Arial", size=10)
            cell.border = Border(top=Side(style="double"))

    # ── Anchos ──
    widths = [5, 12, 6, 7, 10, 8, 12, 32, 13, 13, 10, 13, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A'].width = w

    # Column letters for wider columns
    cols = "ABCDEFGHIJKLMNO"
    for i, w in enumerate(widths):
        ws.column_dimensions[cols[i]].width = w

    # ── Output ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Registro_Ventas_CCPL_{anio}_{mes}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ══════════════════════════════════════════════════════════
# REPORTE DE CAJA (por día)
# ══════════════════════════════════════════════════════════

@router.get("/reporte-caja")
async def reporte_caja(
    fecha: str,
    centro_costo_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    inicio, fin = _rango_dia(fecha)

    query = db.query(Payment).filter(
        Payment.status.in_(["approved", "anulado"]),
        Payment.reviewed_at >= inicio,
        Payment.reviewed_at < fin,
    ).order_by(Payment.reviewed_at)

    pagos = query.all()

    operaciones = []
    for p in pagos:
        col = None
        if p.colegiado_id:
            col = db.query(Colegiado).filter(Colegiado.id == p.colegiado_id).first()

        operaciones.append({
            "id": p.id,
            "hora": p.reviewed_at.strftime("%H:%M") if p.reviewed_at else "",
            "cajero": "",  # TODO: from user session when auth is implemented
            "colegiado": col.apellidos_nombres if col else "",
            "concepto": (p.notes or "").replace("[CAJA] ", "").replace("[ANULADO]", "").strip()[:60],
            "metodo_pago": p.payment_method or "",
            "comprobante": "",  # TODO: from comprobantes_electronicos
            "monto": float(p.amount or 0),
            "status": p.status,
        })

    return {"operaciones": operaciones, "fecha": fecha}


@router.get("/reporte-caja/excel")
async def reporte_caja_excel(
    fecha: str,
    centro_costo_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    data = await reporte_caja(fecha, centro_costo_id, db)
    ops = data["operaciones"]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Caja"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"REPORTE DE CAJA — {fecha}"
    ws["A1"].font = Font(bold=True, size=13, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "COLEGIO DE CONTADORES PÚBLICOS DE LORETO"
    ws["A2"].font = Font(size=11, name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Hora", "Cajero", "Colegiado", "Concepto", "Método", "Comprobante", "Monto", "Estado"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for i, o in enumerate(ops, 1):
        row = i + 4
        vals = [o["hora"], o["cajero"], o["colegiado"], o["concepto"], o["metodo_pago"], o["comprobante"], o["monto"], o["status"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = thin
            if c == 7:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            if o["status"] == "anulado":
                cell.font = Font(color="FF0000", strikethrough=True, name="Arial", size=10)

    # Totales
    if ops:
        tot_row = len(ops) + 5
        ws.cell(row=tot_row, column=6, value="TOTAL").font = Font(bold=True, name="Arial")
        activos = [o for o in ops if o["status"] != "anulado"]
        cell = ws.cell(row=tot_row, column=7, value=sum(o["monto"] for o in activos))
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, name="Arial")
        cell.border = Border(top=Side(style="double"))

    widths = [8, 15, 28, 35, 12, 16, 12, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte_Caja_CCPL_{fecha}.xlsx"'}
    )


# ══════════════════════════════════════════════════════════
# RESUMEN MENSUAL (día por día)
# ══════════════════════════════════════════════════════════

@router.get("/resumen-mensual")
async def resumen_mensual(
    mes: str,
    anio: str,
    db: Session = Depends(get_db),
):
    inicio, fin = _rango_mes(mes, anio)

    pagos = db.query(Payment).filter(
        Payment.status.in_(["approved"]),
        Payment.reviewed_at >= inicio,
        Payment.reviewed_at < fin,
    ).all()

    # Agrupar por día
    dias_dict = {}
    for p in pagos:
        dia = p.reviewed_at.strftime("%d/%m") if p.reviewed_at else "?"
        if dia not in dias_dict:
            dias_dict[dia] = {"dia": dia, "efectivo": 0, "yape_plin": 0, "tarjeta": 0, "transferencia": 0, "total": 0, "egresos": 0, "operaciones": 0}

        monto = float(p.amount or 0)
        met = p.payment_method or ""
        dias_dict[dia]["total"] += monto
        dias_dict[dia]["operaciones"] += 1

        if met == "efectivo":
            dias_dict[dia]["efectivo"] += monto
        elif met in ("yape", "plin"):
            dias_dict[dia]["yape_plin"] += monto
        elif met == "tarjeta":
            dias_dict[dia]["tarjeta"] += monto
        elif met in ("transferencia", "deposito"):
            dias_dict[dia]["transferencia"] += monto

    # TODO: agregar egresos de sesiones_caja

    dias = sorted(dias_dict.values(), key=lambda d: d["dia"])
    return {"dias": dias, "mes": mes, "anio": anio}


# ══════════════════════════════════════════════════════════
# EXPORTAR DÍA (para botón en caja.html)
# ══════════════════════════════════════════════════════════

@router.get("/exportar-dia")
async def exportar_dia(
    fecha: str,
    db: Session = Depends(get_db),
):
    """Redirige al export de reporte-caja/excel"""
    return await reporte_caja_excel(fecha, None, db)
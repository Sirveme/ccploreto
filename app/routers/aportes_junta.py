"""
app/routers/aportes_junta.py
Módulo Aportes a la Junta de Decanos (JDCCPP) — Piezas D, E, J (incremento 1).

Pantallas admin (read/report). Prefix /admin/aportes-junta para evitar colisión
con el módulo de reportes/SUNAT y cualquier ruta /admin/aportes* existente:
  GET  /admin/aportes-junta                      → lista de periodos (E)
  GET  /admin/aportes-junta/config               → configuración read-only (D)
  GET  /admin/aportes-junta/periodo/{periodo_id} → detalle del periodo (E)
  POST /admin/aportes-junta/periodo/{periodo_id}/recalcular → recálculo manual

Acceso: roles admin/decano/tesorero/sote. Org fija = CCPL (1).
Pendiente (incremento 2): F (registro manual), G (carga histórica), H (PDF),
I (Excel), registrar depósito (upload a GCS).
"""

import io
import json
import uuid
from datetime import datetime, date, timezone, timedelta

from fastapi import APIRouter, Depends, Request, HTTPException, UploadFile, File, Form, Body
from fastapi.responses import HTMLResponse, JSONResponse, Response, RedirectResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Member, User, Colegiado
from app.routers.dashboard import get_current_member
from app.utils.templates import templates

TZ_PERU = timezone(timedelta(hours=-5))

router = APIRouter(prefix="/admin/aportes-junta", tags=["aportes-junta"])

ORG_CCPL = 1
_ROLES_APORTES = ("admin", "decano", "tesorero", "sote")
MESES_ES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def require_aportes(current_member: Member = Depends(get_current_member)) -> Member:
    if current_member is None or current_member.role not in _ROLES_APORTES:
        raise HTTPException(status_code=403, detail="Acceso restringido a administración")
    return current_member


def _show_psp_footer(db: Session, current_member: Member) -> bool:
    """Pieza J: el footer de Perú Sistemas Pro aparece solo tras el onboarding
    (cuando el usuario ya cambió su clave inicial)."""
    user = db.query(User).filter(User.id == current_member.user_id).first()
    return bool(user and user.debe_cambiar_clave is False)


# ════════════════════════════════════════════════════════════════
# PIEZA E — LISTA DE PERIODOS
# ════════════════════════════════════════════════════════════════
@router.get("", response_class=HTMLResponse)
async def aportes_lista(
    request: Request,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodos = db.execute(text("""
        SELECT ap.id, ap.anio, ap.mes, ap.estado,
               ap.cantidad_nuevos, ap.monto_nuevos,
               ap.cantidad_habiles, ap.monto_habiles, ap.monto_total,
               ap.cerrado_en,
               dep.id AS deposito_id, dep.numero_voucher, dep.fecha_deposito, dep.monto AS deposito_monto,
               (SELECT COUNT(*) FROM aporte_periodo_alerta a
                 WHERE a.aporte_periodo_id = ap.id AND a.tipo = 'colegiado_sin_pago'
                   AND a.resuelto = FALSE) AS alertas
        FROM aporte_periodos ap
        LEFT JOIN aporte_deposito dep ON dep.aporte_periodo_id = ap.id
        WHERE ap.organizacion_id = :org
        ORDER BY ap.anio DESC, ap.mes DESC
    """), {"org": ORG_CCPL}).fetchall()

    filas = [{
        "id": p.id,
        "periodo_label": f"{MESES_ES[p.mes]} {p.anio}",
        "anio": p.anio, "mes": p.mes,
        "estado": p.estado,
        "cantidad_nuevos": p.cantidad_nuevos or 0,
        "monto_nuevos": float(p.monto_nuevos or 0),
        "cantidad_habiles": p.cantidad_habiles or 0,
        "monto_habiles": float(p.monto_habiles or 0),
        "monto_total": float(p.monto_total or 0),
        "alertas": p.alertas or 0,
        "tiene_deposito": p.deposito_id is not None,
        "numero_voucher": p.numero_voucher,
    } for p in periodos]

    return templates.TemplateResponse("pages/admin/aportes_list.html", {
        "request": request,
        "periodos": filas,
        "show_psp_footer": _show_psp_footer(db, current_member),
    })


# ════════════════════════════════════════════════════════════════
# PIEZA D — CONFIGURACIÓN (read-only)
# ════════════════════════════════════════════════════════════════
@router.get("/config", response_class=HTMLResponse)
async def aportes_config(
    request: Request,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    junta = db.execute(text("""
        SELECT j.id, j.nombre, j.nombre_corto, j.profesion, j.ruc,
               j.banco_destino, j.cuenta_destino, j.moneda, j.titular_cuenta,
               o.codigo_ante_junta
        FROM organizations o
        JOIN juntas j ON j.id = o.junta_id
        WHERE o.id = :org
    """), {"org": ORG_CCPL}).fetchone()

    config = None
    acceso = None
    representante = None
    if junta:
        config = db.execute(text("""
            SELECT * FROM junta_config_aporte
            WHERE junta_id = :jid AND vigencia_hasta IS NULL
            ORDER BY vigencia_desde DESC LIMIT 1
        """), {"jid": junta.id}).fetchone()
        acceso = db.execute(text("""
            SELECT * FROM junta_acceso_config
            WHERE organizacion_id = :o AND junta_id = :j
        """), {"o": ORG_CCPL, "j": junta.id}).fetchone()
        representante = db.execute(text("""
            SELECT u.id, u.public_id, u.name, u.debe_cambiar_clave
            FROM users u JOIN members m ON m.user_id = u.id
            WHERE m.organization_id = :o AND m.role = 'junta_jdccpp'
            ORDER BY u.id LIMIT 1
        """), {"o": ORG_CCPL}).fetchone()

    return templates.TemplateResponse("pages/admin/aportes_config.html", {
        "request": request,
        "junta": junta,
        "config": config,
        "acceso": acceso,
        "representante": representante,
        "show_psp_footer": _show_psp_footer(db, current_member),
    })


@router.put("/config/acceso-junta")
async def config_acceso_junta(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    """Actualiza junta_acceso_config (límite descargas, históricos, ventana)."""
    junta_id, _ = _junta_y_config(db)
    if not junta_id:
        raise HTTPException(400, "Sin Junta asignada")
    max_desc = int(payload.get("max_descargas_por_periodo") or 0)
    ver_hist = bool(payload.get("permite_ver_historicos"))
    sin_restr = bool(payload.get("consulta_sin_restriccion"))
    dias = payload.get("consulta_dias_disponible")
    dias = int(dias) if dias not in (None, "", False) else None
    db.execute(text("""
        INSERT INTO junta_acceso_config (
            organizacion_id, junta_id, consulta_sin_restriccion, consulta_dias_disponible,
            max_descargas_por_periodo, permite_ver_historicos, updated_at, updated_by_user_id
        ) VALUES (:o, :j, :sr, :dias, :md, :vh, NOW(), :uid)
        ON CONFLICT (organizacion_id, junta_id) DO UPDATE SET
            consulta_sin_restriccion = EXCLUDED.consulta_sin_restriccion,
            consulta_dias_disponible = EXCLUDED.consulta_dias_disponible,
            max_descargas_por_periodo = EXCLUDED.max_descargas_por_periodo,
            permite_ver_historicos = EXCLUDED.permite_ver_historicos,
            updated_at = NOW(), updated_by_user_id = EXCLUDED.updated_by_user_id
    """), {"o": ORG_CCPL, "j": junta_id, "sr": sin_restr, "dias": dias,
           "md": max_desc, "vh": ver_hist, "uid": current_member.user_id})
    db.commit()
    return JSONResponse({"ok": True})


@router.put("/config/representante-junta")
async def config_representante(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    """Actualiza DNI/nombre del representante JDCCPP. Al cambiar el DNI se recalcula
    el access_code (Argon2 del nuevo DNI = clave inicial) y se fuerza cambio de clave."""
    rep = db.execute(text("""
        SELECT u.id, u.public_id FROM users u JOIN members m ON m.user_id = u.id
        WHERE m.organization_id = :o AND m.role = 'junta_jdccpp' ORDER BY u.id LIMIT 1
    """), {"o": ORG_CCPL}).fetchone()
    if not rep:
        raise HTTPException(404, "Representante JDCCPP no encontrado")

    nuevo_dni = (payload.get("public_id") or "").strip()
    nuevo_nombre = (payload.get("name") or "").strip()
    forzar = bool(payload.get("forzar_reset"))

    from app.utils.security import get_password_hash

    fields = {}
    if nuevo_nombre:
        fields["name"] = nuevo_nombre
    dni_cambia = nuevo_dni and nuevo_dni != rep.public_id
    if dni_cambia:
        # unicidad
        dup = db.execute(text("SELECT 1 FROM users WHERE public_id = :d AND id <> :id"),
                         {"d": nuevo_dni, "id": rep.id}).fetchone()
        if dup:
            raise HTTPException(409, "Ya existe un usuario con ese DNI")
        fields["public_id"] = nuevo_dni
        fields["access_code"] = get_password_hash(nuevo_dni)
        fields["debe_cambiar_clave"] = True
    elif forzar:
        fields["access_code"] = get_password_hash(rep.public_id)
        fields["debe_cambiar_clave"] = True

    if not fields:
        return JSONResponse({"ok": True, "sin_cambios": True})

    set_clause = ", ".join(f"{k} = :{k}" for k in fields)
    fields["id"] = rep.id
    db.execute(text(f"UPDATE users SET {set_clause} WHERE id = :id"), fields)
    db.commit()
    return JSONResponse({"ok": True, "dni_cambiado": dni_cambia})


# ════════════════════════════════════════════════════════════════
# PIEZA E — DETALLE DEL PERIODO
# ════════════════════════════════════════════════════════════════
@router.get("/periodo/{periodo_id}", response_class=HTMLResponse)
async def aportes_detalle(
    periodo_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo = db.execute(text("""
        SELECT ap.*, dep.id AS deposito_id, dep.numero_voucher, dep.fecha_deposito,
               dep.monto AS deposito_monto, dep.banco_emisor, dep.evidencia_url
        FROM aporte_periodos ap
        LEFT JOIN aporte_deposito dep ON dep.aporte_periodo_id = ap.id
        WHERE ap.id = :pid AND ap.organizacion_id = :org
    """), {"pid": periodo_id, "org": ORG_CCPL}).fetchone()

    if not periodo:
        raise HTTPException(status_code=404, detail="Periodo no encontrado")

    nuevos = db.execute(text("""
        SELECT codigo_matricula, apellidos_nombres, dni,
               fecha_pago_der_col, fecha_colegiatura, monto_aporte,
               fuente_registro
        FROM aporte_detalle_nuevos
        WHERE aporte_periodo_id = :pid
        ORDER BY codigo_matricula NULLS LAST, apellidos_nombres
    """), {"pid": periodo_id}).fetchall()

    alertas = db.execute(text("""
        SELECT colegiado_id, mensaje, created_at
        FROM aporte_periodo_alerta
        WHERE aporte_periodo_id = :pid AND tipo = 'colegiado_sin_pago' AND resuelto = FALSE
        ORDER BY created_at
    """), {"pid": periodo_id}).fetchall()

    ctx = {
        "request": request,
        "periodo": periodo,
        "periodo_label": f"{MESES_ES[periodo.mes]} {periodo.anio}",
        "nuevos": nuevos,
        "alertas": alertas,
        "show_psp_footer": _show_psp_footer(db, current_member),
    }
    return templates.TemplateResponse("pages/admin/aportes_detalle.html", ctx)


# ════════════════════════════════════════════════════════════════
# Recálculo manual (solo periodos abiertos)
# ════════════════════════════════════════════════════════════════
@router.post("/periodo/{periodo_id}/recalcular")
async def aportes_recalcular(
    periodo_id: int,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo = db.execute(text("""
        SELECT estado, anio, mes FROM aporte_periodos
        WHERE id = :pid AND organizacion_id = :org
    """), {"pid": periodo_id, "org": ORG_CCPL}).fetchone()

    if not periodo:
        raise HTTPException(status_code=404, detail="Periodo no encontrado")
    if periodo.estado == "cerrado":
        return JSONResponse({"ok": False, "error": "El periodo está cerrado (inmutable)."},
                            status_code=400)

    from app.services.aportes_junta_service import calcular_periodo_actual
    from datetime import datetime
    from app.services.aportes_junta_service import TZ_PERU
    ahora = datetime.now(TZ_PERU)
    if (periodo.anio, periodo.mes) != (ahora.year, ahora.month):
        return JSONResponse(
            {"ok": False, "error": "Solo se recalcula el periodo del mes en curso. "
             "Los anteriores se gestionan por carga manual."},
            status_code=400,
        )

    result = calcular_periodo_actual(db, organizacion_id=ORG_CCPL)
    if not result:
        return JSONResponse({"ok": False, "error": "No se pudo recalcular."}, status_code=400)
    return JSONResponse({"ok": True, "result": {
        "cantidad_nuevos": result["cantidad_nuevos"],
        "cantidad_habiles": result["cantidad_habiles"],
        "monto_total": result["monto_total"],
        "pendientes_registro": result["pendientes_registro"],
    }})


# ════════════════════════════════════════════════════════════════
# HELPERS (Incremento 2)
# ════════════════════════════════════════════════════════════════
def _junta_y_config(db: Session):
    org = db.execute(text("SELECT junta_id FROM organizations WHERE id = :o"),
                     {"o": ORG_CCPL}).fetchone()
    if not org or not org.junta_id:
        return None, None
    cfg = db.execute(text("""
        SELECT * FROM junta_config_aporte
        WHERE junta_id = :j AND vigencia_desde <= :hoy
          AND (vigencia_hasta IS NULL OR vigencia_hasta >= :hoy)
        ORDER BY vigencia_desde DESC LIMIT 1
    """), {"j": org.junta_id, "hoy": datetime.now(TZ_PERU).date()}).fetchone()
    return org.junta_id, cfg


def _get_or_create_periodo(db: Session, junta_id: int, anio: int, mes: int):
    return db.execute(text("""
        INSERT INTO aporte_periodos (organizacion_id, junta_id, anio, mes, estado, created_at, updated_at)
        VALUES (:org, :j, :a, :m, 'abierto', NOW(), NOW())
        ON CONFLICT (organizacion_id, anio, mes) DO UPDATE SET updated_at = NOW()
        RETURNING id, estado
    """), {"org": ORG_CCPL, "j": junta_id, "a": anio, "m": mes}).fetchone()


def _subir_gcs(contenido: bytes, filename: str, content_type: str, anio: int, mes: int):
    """Sube evidencia/voucher a GCS. Retorna (url, error)."""
    from app.utils.gcs import _get_client, BUCKET_NAME
    ext = (filename.rsplit('.', 1)[-1].lower() if filename and '.' in filename else 'bin')
    if ext not in ('pdf', 'jpg', 'jpeg', 'png', 'webp'):
        return None, "Formato no permitido (pdf/jpg/png/webp)"
    client = _get_client()
    if not client:
        return None, "GCS no configurado"
    nombre = f"{uuid.uuid4().hex[:12]}.{ext}"
    blob_path = f"{ORG_CCPL}/aportes/{anio}/{mes:02d}/{nombre}"
    blob = client.bucket(BUCKET_NAME).blob(blob_path)
    blob.upload_from_string(contenido, content_type=content_type or "application/octet-stream")
    return f"https://storage.googleapis.com/{BUCKET_NAME}/{blob_path}", None


def _recalcular_totales(db: Session, periodo_id: int, monto_por_habil: float):
    """Recalcula nuevos (desde el detalle) + hábiles (live, fin de mes del periodo)."""
    from calendar import monthrange
    per = db.execute(text("SELECT anio, mes FROM aporte_periodos WHERE id = :p"),
                     {"p": periodo_id}).fetchone()
    _, ud = monthrange(per.anio, per.mes)
    corte = datetime(per.anio, per.mes, ud, 23, 59, 59, tzinfo=TZ_PERU)
    tot = db.execute(text("""
        SELECT COUNT(*) AS c, COALESCE(SUM(monto_aporte), 0) AS s
        FROM aporte_detalle_nuevos WHERE aporte_periodo_id = :p
    """), {"p": periodo_id}).fetchone()
    hab = db.execute(text("""
        SELECT COUNT(*) AS c FROM colegiados
        WHERE organization_id = :o AND condicion = 'habil' AND habilidad_vence >= :corte
          AND COALESCE(aporta_jdccpp, TRUE) = TRUE
    """), {"o": ORG_CCPL, "corte": corte}).fetchone()
    cn, mn = tot.c, float(tot.s)
    ch = hab.c or 0
    mh = ch * float(monto_por_habil)
    db.execute(text("""
        UPDATE aporte_periodos SET cantidad_nuevos = :cn, monto_nuevos = :mn,
               cantidad_habiles = :ch, monto_habiles = :mh, monto_total = :mt, updated_at = NOW()
        WHERE id = :p
    """), {"cn": cn, "mn": mn, "ch": ch, "mh": mh, "mt": mn + mh, "p": periodo_id})


def _fetch_reporte(db: Session, periodo_id: int):
    """Datos completos para PDF/Excel."""
    periodo = db.execute(text("""
        SELECT ap.*, dep.numero_voucher, dep.fecha_deposito, dep.monto AS deposito_monto,
               dep.banco_emisor
        FROM aporte_periodos ap
        LEFT JOIN aporte_deposito dep ON dep.aporte_periodo_id = ap.id
        WHERE ap.id = :pid AND ap.organizacion_id = :org
    """), {"pid": periodo_id, "org": ORG_CCPL}).fetchone()
    if not periodo:
        return None, None
    nuevos = db.execute(text("""
        SELECT codigo_matricula, apellidos_nombres, dni, fecha_pago_der_col, monto_aporte
        FROM aporte_detalle_nuevos WHERE aporte_periodo_id = :pid
        ORDER BY codigo_matricula NULLS LAST, apellidos_nombres
    """), {"pid": periodo_id}).fetchall()
    return periodo, nuevos


# ════════════════════════════════════════════════════════════════
# Búsqueda de colegiado (para F y G)
# ════════════════════════════════════════════════════════════════
@router.get("/buscar-colegiado")
async def buscar_colegiado(
    q: str,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    q = (q or "").strip()
    if len(q) < 2:
        return JSONResponse({"resultados": []})
    filas = db.execute(text("""
        SELECT id, codigo_matricula, apellidos_nombres, dni, fecha_colegiatura
        FROM colegiados
        WHERE organization_id = :org
          AND (codigo_matricula ILIKE :q OR dni ILIKE :q OR apellidos_nombres ILIKE :qn)
        ORDER BY apellidos_nombres LIMIT 15
    """), {"org": ORG_CCPL, "q": f"%{q}%", "qn": f"%{q}%"}).fetchall()
    return JSONResponse({"resultados": [{
        "id": r.id, "codigo_matricula": r.codigo_matricula,
        "apellidos_nombres": r.apellidos_nombres, "dni": r.dni,
        "fecha_colegiatura": r.fecha_colegiatura.isoformat() if r.fecha_colegiatura else None,
    } for r in filas]})


# ════════════════════════════════════════════════════════════════
# PIEZA F — REGISTRO MANUAL DE APORTE POR NUEVO COLEGIADO
# ════════════════════════════════════════════════════════════════
@router.get("/nuevo", response_class=HTMLResponse)
async def aportes_nuevo_form(
    request: Request,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    _, cfg = _junta_y_config(db)
    return templates.TemplateResponse("pages/admin/aportes_nuevo.html", {
        "request": request,
        "monto_aporte": float(cfg.monto_por_nuevo) if cfg else 0.0,
        "hoy": datetime.now(TZ_PERU).date().isoformat(),
        "show_psp_footer": _show_psp_footer(db, current_member),
    })


@router.post("/nuevo")
async def aportes_nuevo_submit(
    colegiado_id: int = Form(...),
    fecha_pago: str = Form(...),
    monto_pagado: float = Form(None),
    forma_pago: str = Form(None),
    numero_operacion: str = Form(None),
    observaciones: str = Form(None),
    evidencia: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    junta_id, cfg = _junta_y_config(db)
    if not cfg:
        raise HTTPException(400, "Sin configuración de aporte vigente")
    try:
        fpago = datetime.strptime(fecha_pago, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Fecha de pago inválida")

    col = db.execute(text("""
        SELECT id, codigo_matricula, apellidos_nombres, dni, fecha_colegiatura
        FROM colegiados WHERE id = :cid AND organization_id = :org
    """), {"cid": colegiado_id, "org": ORG_CCPL}).fetchone()
    if not col:
        raise HTTPException(404, "Colegiado no encontrado")

    periodo = _get_or_create_periodo(db, junta_id, fpago.year, fpago.month)
    if periodo.estado == "cerrado":
        raise HTTPException(400, "El periodo de esa fecha está cerrado (inmutable)")

    evidencia_url = None
    if evidencia and evidencia.filename:
        contenido = await evidencia.read()
        evidencia_url, err = _subir_gcs(contenido, evidencia.filename,
                                        evidencia.content_type, fpago.year, fpago.month)
        if err:
            raise HTTPException(400, f"Evidencia: {err}")

    obs = (observaciones or "").strip()
    if forma_pago:
        obs = (f"Forma de pago: {forma_pago}. " + obs).strip()
    if numero_operacion:
        obs = (obs + f" N° op: {numero_operacion}").strip()

    db.execute(text("""
        INSERT INTO aporte_detalle_nuevos (
            aporte_periodo_id, colegiado_id, payment_id, codigo_matricula,
            apellidos_nombres, dni, fecha_pago_der_col, fecha_colegiatura,
            monto_pagado, monto_aporte, fuente_registro, observaciones,
            evidencia_url, registrado_por_user_id, codigo_lote, created_at
        ) VALUES (
            :pid, :cid, NULL, :mat, :nom, :dni, :fpago, :fcol,
            :mpago, :maporte, 'manual_caja', :obs, :evi, :uid, :lote, NOW()
        )
        ON CONFLICT (aporte_periodo_id, colegiado_id) DO UPDATE SET
            fecha_pago_der_col = EXCLUDED.fecha_pago_der_col,
            monto_pagado = EXCLUDED.monto_pagado,
            monto_aporte = EXCLUDED.monto_aporte,
            fuente_registro = 'manual_caja',
            observaciones = EXCLUDED.observaciones,
            evidencia_url = COALESCE(EXCLUDED.evidencia_url, aporte_detalle_nuevos.evidencia_url),
            registrado_por_user_id = EXCLUDED.registrado_por_user_id
    """), {
        "pid": periodo.id, "cid": col.id, "mat": col.codigo_matricula,
        "nom": col.apellidos_nombres, "dni": col.dni, "fpago": fpago,
        "fcol": (col.fecha_colegiatura.date() if hasattr(col.fecha_colegiatura, "date") else col.fecha_colegiatura) if col.fecha_colegiatura else None,
        "mpago": monto_pagado, "maporte": float(cfg.monto_por_nuevo),
        "obs": obs or None, "evi": evidencia_url, "uid": current_member.user_id,
        "lote": f"LOTE-{fpago.year}{fpago.month:02d}-{ORG_CCPL:03d}",
    })

    # Resolver alerta de "colegiado_sin_pago" si existía
    db.execute(text("""
        UPDATE aporte_periodo_alerta SET resuelto = TRUE, resuelto_at = NOW()
        WHERE aporte_periodo_id = :pid AND colegiado_id = :cid
          AND tipo = 'colegiado_sin_pago' AND resuelto = FALSE
    """), {"pid": periodo.id, "cid": col.id})

    _recalcular_totales(db, periodo.id, cfg.monto_por_habil)
    db.commit()
    return RedirectResponse(url=f"/admin/aportes-junta/periodo/{periodo.id}", status_code=303)


# ════════════════════════════════════════════════════════════════
# DEPÓSITO — registrar voucher BBVA (upload GCS)
# ════════════════════════════════════════════════════════════════
@router.post("/periodo/{periodo_id}/deposito")
async def aportes_deposito(
    periodo_id: int,
    monto: float = Form(...),
    fecha_deposito: str = Form(...),
    numero_voucher: str = Form(None),
    banco_emisor: str = Form(None),
    cuenta_origen: str = Form(None),
    observaciones: str = Form(None),
    evidencia: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo = db.execute(text("""
        SELECT id, anio, mes FROM aporte_periodos
        WHERE id = :pid AND organizacion_id = :org
    """), {"pid": periodo_id, "org": ORG_CCPL}).fetchone()
    if not periodo:
        raise HTTPException(404, "Periodo no encontrado")
    try:
        fdep = datetime.strptime(fecha_deposito, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Fecha de depósito inválida")

    evidencia_url = None
    if evidencia and evidencia.filename:
        contenido = await evidencia.read()
        evidencia_url, err = _subir_gcs(contenido, evidencia.filename,
                                        evidencia.content_type, periodo.anio, periodo.mes)
        if err:
            raise HTTPException(400, f"Voucher: {err}")

    db.execute(text("""
        INSERT INTO aporte_deposito (
            aporte_periodo_id, monto, fecha_deposito, numero_voucher, banco_emisor,
            cuenta_origen, evidencia_url, observaciones, registrado_por_user_id, created_at, updated_at
        ) VALUES (
            :pid, :monto, :fdep, :voucher, :banco, :cuenta, :evi, :obs, :uid, NOW(), NOW()
        )
        ON CONFLICT (aporte_periodo_id) DO UPDATE SET
            monto = EXCLUDED.monto, fecha_deposito = EXCLUDED.fecha_deposito,
            numero_voucher = EXCLUDED.numero_voucher, banco_emisor = EXCLUDED.banco_emisor,
            cuenta_origen = EXCLUDED.cuenta_origen,
            evidencia_url = COALESCE(EXCLUDED.evidencia_url, aporte_deposito.evidencia_url),
            observaciones = EXCLUDED.observaciones, updated_at = NOW()
    """), {
        "pid": periodo_id, "monto": monto, "fdep": fdep, "voucher": numero_voucher,
        "banco": banco_emisor, "cuenta": cuenta_origen, "evi": evidencia_url,
        "obs": observaciones, "uid": current_member.user_id,
    })
    db.commit()
    return RedirectResponse(url=f"/admin/aportes-junta/periodo/{periodo_id}", status_code=303)


# ════════════════════════════════════════════════════════════════
# PIEZA G — CARGA HISTÓRICA
# ════════════════════════════════════════════════════════════════
@router.get("/carga-historica", response_class=HTMLResponse)
async def aportes_carga_form(
    request: Request,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    _, cfg = _junta_y_config(db)
    return templates.TemplateResponse("pages/admin/aportes_carga_historica.html", {
        "request": request,
        "monto_aporte": float(cfg.monto_por_nuevo) if cfg else 0.0,
        "monto_habil": float(cfg.monto_por_habil) if cfg else 0.0,
        "meses": list(enumerate(MESES_ES))[1:],
        "show_psp_footer": _show_psp_footer(db, current_member),
    })


@router.post("/carga-historica")
async def aportes_carga_submit(
    anio: int = Form(...),
    mes: int = Form(...),
    cantidad_habiles: int = Form(...),
    nominal_json: str = Form(...),
    dep_monto: float = Form(None),
    dep_fecha: str = Form(None),
    dep_voucher: str = Form(None),
    dep_banco: str = Form(None),
    voucher: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    junta_id, cfg = _junta_y_config(db)
    if not cfg:
        raise HTTPException(400, "Sin configuración de aporte vigente")
    try:
        nominal = json.loads(nominal_json or "[]")
    except json.JSONDecodeError:
        raise HTTPException(400, "Lista nominal inválida")
    if not isinstance(nominal, list) or not nominal:
        raise HTTPException(400, "Debe ingresar al menos un colegiado nuevo")

    periodo = _get_or_create_periodo(db, junta_id, anio, mes)
    if periodo.estado == "cerrado":
        raise HTTPException(400, "Ese periodo ya está cerrado. No se puede recargar.")

    maporte = float(cfg.monto_por_nuevo)
    lote = f"LOTE-{anio}{mes:02d}-{ORG_CCPL:03d}"

    # Re-carga idempotente: limpiar lo histórico previo de este periodo.
    db.execute(text("""
        DELETE FROM aporte_detalle_nuevos
        WHERE aporte_periodo_id = :pid AND fuente_registro = 'carga_historica'
    """), {"pid": periodo.id})

    for row in nominal:
        nom = (row.get("apellidos_nombres") or "").strip()
        if not nom:
            continue
        fpago = None
        if row.get("fecha_pago"):
            try:
                fpago = datetime.strptime(row["fecha_pago"], "%Y-%m-%d").date()
            except ValueError:
                fpago = None
        db.execute(text("""
            INSERT INTO aporte_detalle_nuevos (
                aporte_periodo_id, colegiado_id, payment_id, codigo_matricula,
                apellidos_nombres, dni, fecha_pago_der_col, monto_pagado, monto_aporte,
                fuente_registro, registrado_por_user_id, codigo_lote, created_at
            ) VALUES (
                :pid, :cid, NULL, :mat, :nom, :dni, :fpago, :mpago, :maporte,
                'carga_historica', :uid, :lote, NOW()
            )
        """), {
            "pid": periodo.id, "cid": row.get("colegiado_id") or None,
            "mat": (row.get("codigo_matricula") or None),
            "nom": nom, "dni": (row.get("dni") or None), "fpago": fpago,
            "mpago": row.get("monto_pagado"), "maporte": maporte,
            "uid": current_member.user_id, "lote": lote,
        })

    # Totales: nuevos desde el detalle; hábiles desde el formulario (histórico).
    tot = db.execute(text("""
        SELECT COUNT(*) AS c, COALESCE(SUM(monto_aporte), 0) AS s
        FROM aporte_detalle_nuevos WHERE aporte_periodo_id = :p
    """), {"p": periodo.id}).fetchone()
    cn, mn = tot.c, float(tot.s)
    ch = int(cantidad_habiles)
    mh = ch * float(cfg.monto_por_habil)

    # Depósito (opcional) + voucher GCS
    if dep_monto is not None and dep_fecha:
        try:
            fdep = datetime.strptime(dep_fecha, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, "Fecha de depósito inválida")
        evidencia_url = None
        if voucher and voucher.filename:
            contenido = await voucher.read()
            evidencia_url, err = _subir_gcs(contenido, voucher.filename,
                                            voucher.content_type, anio, mes)
            if err:
                raise HTTPException(400, f"Voucher: {err}")
        db.execute(text("""
            INSERT INTO aporte_deposito (
                aporte_periodo_id, monto, fecha_deposito, numero_voucher, banco_emisor,
                evidencia_url, registrado_por_user_id, created_at, updated_at
            ) VALUES (:pid, :monto, :fdep, :voucher, :banco, :evi, :uid, NOW(), NOW())
            ON CONFLICT (aporte_periodo_id) DO UPDATE SET
                monto = EXCLUDED.monto, fecha_deposito = EXCLUDED.fecha_deposito,
                numero_voucher = EXCLUDED.numero_voucher, banco_emisor = EXCLUDED.banco_emisor,
                evidencia_url = COALESCE(EXCLUDED.evidencia_url, aporte_deposito.evidencia_url),
                updated_at = NOW()
        """), {"pid": periodo.id, "monto": dep_monto, "fdep": fdep, "voucher": dep_voucher,
               "banco": dep_banco, "evi": evidencia_url, "uid": current_member.user_id})

    # Cerrar el periodo con snapshot de la config aplicada (inmutable).
    db.execute(text("""
        UPDATE aporte_periodos SET
            cantidad_nuevos = :cn, monto_nuevos = :mn,
            cantidad_habiles = :ch, monto_habiles = :mh, monto_total = :mt,
            estado = 'cerrado', cerrado_en = NOW(), cerrado_por = :uid,
            uit_aplicada = :uit, monto_por_nuevo_aplicado = :mpn, monto_por_habil_aplicado = :mph,
            pct_nuevo_aplicado = :pn, pct_habil_aplicado = :ph, base_cuota_aplicada = :bc,
            codigo_lote = :lote, updated_at = NOW()
        WHERE id = :pid
    """), {
        "cn": cn, "mn": mn, "ch": ch, "mh": mh, "mt": mn + mh, "uid": str(current_member.user_id),
        "uit": cfg.base_uit, "mpn": cfg.monto_por_nuevo, "mph": cfg.monto_por_habil,
        "pn": cfg.pct_sobre_uit_nuevo, "ph": cfg.pct_sobre_cuota_habil,
        "bc": cfg.base_cuota_ordinaria, "lote": lote, "pid": periodo.id,
    })
    db.execute(text("""
        INSERT INTO aporte_periodo_log (
            aporte_periodo_id, cantidad_nuevos, monto_nuevos, cantidad_habiles,
            monto_habiles, monto_total, evento, detalle
        ) VALUES (:pid, :cn, :mn, :ch, :mh, :mt, 'carga_historica', :det)
    """), {"pid": periodo.id, "cn": cn, "mn": mn, "ch": ch, "mh": mh, "mt": mn + mh,
           "det": f"Carga histórica {anio}-{mes:02d}: {cn} nuevos + {ch} hábiles"})

    db.commit()
    return RedirectResponse(url=f"/admin/aportes-junta/periodo/{periodo.id}", status_code=303)


# ════════════════════════════════════════════════════════════════
# PIEZA H — EXPORT PDF PLANILLA OFICIAL
# ════════════════════════════════════════════════════════════════
@router.get("/periodo/{periodo_id}/pdf")
async def aportes_pdf(
    periodo_id: int,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo, nuevos = _fetch_reporte(db, periodo_id)
    if not periodo:
        raise HTTPException(404, "Periodo no encontrado")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=14)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=11, alignment=1)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.grey, alignment=1)
    periodo_label = f"{MESES_ES[periodo.mes]} {periodo.anio}"

    el = []
    el.append(Paragraph("COLEGIO DE CONTADORES PÚBLICOS DE LORETO", h))
    el.append(Paragraph(f"Resumen de depósito mensual a JDCCPP — {periodo_label.upper()}", sub))
    el.append(Spacer(1, 10 * mm))

    resumen = [
        ["TOTAL CUOTAS ORDINARIAS MIEMBROS HÁBILES", f"S/ {float(periodo.monto_habiles or 0):,.2f}"],
        ["TOTAL NUEVOS COLEGIADOS", f"S/ {float(periodo.monto_nuevos or 0):,.2f}"],
        [f"TOTAL A DEPOSITAR A JDCCPP — {periodo_label.upper()}", f"S/ {float(periodo.monto_total or 0):,.2f}"],
    ]
    t = Table(resumen, colWidths=[120 * mm, 50 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, -1), (-1, -1), 1, colors.black),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    el.append(t)
    el.append(Spacer(1, 8 * mm))

    el.append(Paragraph("DETALLE DE NUEVOS COLEGIADOS", styles["Heading4"]))
    data = [["N°", "Matrícula", "Apellidos y Nombres", "DNI", "F. Pago", "Monto"]]
    for i, n in enumerate(nuevos, 1):
        data.append([
            str(i), n.codigo_matricula or "—", n.apellidos_nombres, n.dni or "—",
            n.fecha_pago_der_col.strftime("%d/%m/%Y") if n.fecha_pago_der_col else "—",
            f"S/ {float(n.monto_aporte or 0):,.2f}",
        ])
    td = Table(data, colWidths=[10 * mm, 22 * mm, 78 * mm, 24 * mm, 20 * mm, 24 * mm], repeatRows=1)
    td.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ALIGN", (0, 0), (0, -1), "CENTER"), ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    el.append(td)
    el.append(Spacer(1, 16 * mm))

    if periodo.numero_voucher:
        el.append(Paragraph(
            f"Depósito: S/ {float(periodo.deposito_monto or 0):,.2f} · Voucher {periodo.numero_voucher}"
            f" · {periodo.banco_emisor or ''}"
            + (f" · {periodo.fecha_deposito.strftime('%d/%m/%Y')}" if periodo.fecha_deposito else ""),
            styles["Normal"]))
        el.append(Spacer(1, 10 * mm))

    el.append(Paragraph("____________________________<br/>Administrador CCPL", styles["Normal"]))

    if _show_psp_footer(db, current_member):
        el.append(Spacer(1, 12 * mm))
        el.append(Paragraph(
            "Sistema desarrollado por Perú Sistemas Pro · perusistemas.pro · WhatsApp +51 967 317 946",
            small))

    doc.build(el)
    buf.seek(0)
    fname = f"aporte_jdccpp_{periodo.anio}_{periodo.mes:02d}.pdf"
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{fname}"'})


# ════════════════════════════════════════════════════════════════
# PIEZA I — EXPORT EXCEL (3 hojas)
# ════════════════════════════════════════════════════════════════
@router.get("/periodo/{periodo_id}/excel")
async def aportes_excel(
    periodo_id: int,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo, nuevos = _fetch_reporte(db, periodo_id)
    if not periodo:
        raise HTTPException(404, "Periodo no encontrado")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    periodo_label = f"{MESES_ES[periodo.mes]} {periodo.anio}"
    azul = PatternFill("solid", fgColor="1E3A5F")
    blanco_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"
    ws1["A1"] = f"Aporte a JDCCPP — {periodo_label}"; ws1["A1"].font = Font(bold=True, size=14)
    rows = [
        ("", ""),
        ("Cuotas ordinarias hábiles", float(periodo.monto_habiles or 0)),
        (f"   ({periodo.cantidad_habiles or 0} hábiles)", ""),
        ("Nuevos colegiados", float(periodo.monto_nuevos or 0)),
        (f"   ({periodo.cantidad_nuevos or 0} nuevos)", ""),
        ("TOTAL A DEPOSITAR", float(periodo.monto_total or 0)),
        ("", ""),
        ("Estado", periodo.estado),
        ("Voucher", periodo.numero_voucher or "—"),
        ("Depósito S/", float(periodo.deposito_monto or 0)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws1[f"A{i}"] = k; ws1[f"B{i}"] = v
    ws1["A8"].font = bold; ws1["B8"].font = bold
    ws1.column_dimensions["A"].width = 34; ws1.column_dimensions["B"].width = 18

    ws2 = wb.create_sheet("Detalle Nuevos")
    headers = ["N°", "Matrícula", "Apellidos y Nombres", "DNI", "Fecha Pago", "Monto Aporte"]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c); cell.fill = azul; cell.font = blanco_bold
        cell.alignment = Alignment(horizontal="center")
    for i, n in enumerate(nuevos, 1):
        ws2.append([
            i, n.codigo_matricula or "—", n.apellidos_nombres, n.dni or "—",
            n.fecha_pago_der_col.strftime("%d/%m/%Y") if n.fecha_pago_der_col else "—",
            float(n.monto_aporte or 0),
        ])
    for col, w in zip("ABCDEF", [6, 14, 44, 14, 14, 14]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Metadata")
    meta = [
        ("Periodo", periodo_label),
        ("Generado", datetime.now(TZ_PERU).strftime("%d/%m/%Y %H:%M")),
        ("Total nuevos", periodo.cantidad_nuevos or 0),
        ("Total hábiles", periodo.cantidad_habiles or 0),
        ("UIT aplicada", float(periodo.uit_aplicada or 0)),
        ("Total a depositar", float(periodo.monto_total or 0)),
        ("Voucher", periodo.numero_voucher or "—"),
        ("Marco normativo", "Acuerdo institucional JDCCPP"),
    ]
    for k, v in meta:
        ws3.append([k, v])
    if _show_psp_footer(db, current_member):
        ws3.append(["", ""])
        ws3.append(["Sistema", "Perú Sistemas Pro · perusistemas.pro · WhatsApp +51 967 317 946"])
    ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 48

    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f"aporte_jdccpp_{periodo.anio}_{periodo.mes:02d}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ════════════════════════════════════════════════════════════════
# PIEZA C — DASHBOARD DEL ADMINISTRADOR
# ════════════════════════════════════════════════════════════════
@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard_admin(
    request: Request,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    _, cfg = _junta_y_config(db)
    monto_habil = float(cfg.monto_por_habil) if cfg else 0.0
    monto_nuevo = float(cfg.monto_por_nuevo) if cfg else 0.0

    # Resumen del padrón por condición (dinámico) + split aportantes/Past Decano.
    filas = db.execute(text("""
        SELECT condicion,
               COUNT(*) AS total,
               COUNT(*) FILTER (WHERE COALESCE(aporta_jdccpp, TRUE) = TRUE) AS aportan
        FROM colegiados WHERE organization_id = :org
        GROUP BY condicion ORDER BY total DESC
    """), {"org": ORG_CCPL}).fetchall()

    cond = {}
    total_padron = 0
    for f in filas:
        c = f.condicion or "(sin)"
        cond[c] = {"total": f.total, "aportan": f.aportan}
        total_padron += f.total
    habiles_aportantes = cond.get("habil", {}).get("aportan", 0)
    past_decanos = cond.get("habil", {}).get("total", 0) - habiles_aportantes

    # Periodo en curso: crear/recalcular si no existe.
    ahora = datetime.now(TZ_PERU)
    periodo = db.execute(text("""
        SELECT * FROM aporte_periodos WHERE organizacion_id = :org AND anio = :a AND mes = :m
    """), {"org": ORG_CCPL, "a": ahora.year, "m": ahora.month}).fetchone()
    if not periodo:
        try:
            from app.services.aportes_junta_service import calcular_periodo_actual
            calcular_periodo_actual(db, ORG_CCPL)
            periodo = db.execute(text("""
                SELECT * FROM aporte_periodos WHERE organizacion_id = :org AND anio = :a AND mes = :m
            """), {"org": ORG_CCPL, "a": ahora.year, "m": ahora.month}).fetchone()
        except Exception:
            periodo = None

    nuevos = []
    if periodo:
        nuevos = db.execute(text("""
            SELECT dni, apellidos_nombres, codigo_matricula, fecha_colegiatura,
                   fecha_pago_der_col, monto_aporte
            FROM aporte_detalle_nuevos WHERE aporte_periodo_id = :p
            ORDER BY codigo_matricula NULLS LAST, apellidos_nombres
        """), {"p": periodo.id}).fetchall()

    cantidad_nuevos = (periodo.cantidad_nuevos if periodo else 0) or 0
    est_habiles = habiles_aportantes * monto_habil
    est_nuevos = float(periodo.monto_nuevos) if periodo and periodo.monto_nuevos else 0.0

    cards = [
        {"label": "Hábiles aportantes", "icon": "ph-check-circle", "color": "#10b981",
         "total": habiles_aportantes, "monto": est_habiles},
        {"label": "Vitalicios", "icon": "ph-medal", "color": "#3b82f6",
         "total": cond.get("vitalicio", {}).get("total", 0), "monto": None},
        {"label": "Inhábiles", "icon": "ph-x-circle", "color": "#f59e0b",
         "total": cond.get("inhabil", {}).get("total", 0), "monto": None},
        {"label": "Nuevos del mes", "icon": "ph-user-plus", "color": "#8b5cf6",
         "total": cantidad_nuevos, "monto": est_nuevos},
        {"label": "Past Decanos", "icon": "ph-crown", "color": "#ef4444",
         "total": past_decanos, "monto": None},
        {"label": "Candidatos a retiro", "icon": "ph-warning", "color": "#fbbf24",
         "total": cond.get("candidato_retiro", {}).get("total", 0), "monto": None},
    ]

    return templates.TemplateResponse("pages/admin/aportes_dashboard.html", {
        "request": request,
        "cards": cards,
        "total_padron": total_padron,
        "monto_habil": monto_habil,
        "monto_nuevo": monto_nuevo,
        "periodo": periodo,
        "periodo_label": f"{MESES_ES[ahora.month]} {ahora.year}",
        "est_total": est_habiles + est_nuevos,
        "nuevos": nuevos,
        "show_psp_footer": _show_psp_footer(db, current_member),
    })


# ════════════════════════════════════════════════════════════════
# PADRÓN — datos paginados (AJAX) + edición inline (condición / aporta)
# ════════════════════════════════════════════════════════════════
_CONDICIONES_VALIDAS = ("habil", "inhabil", "vitalicio", "candidato_retiro")


@router.get("/padron-data")
async def padron_data(
    condicion: str = None,
    aporta: str = None,
    q: str = None,
    page: int = 1,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    per_page = 50
    page = max(1, page)
    where = ["organization_id = :org"]
    params = {"org": ORG_CCPL, "lim": per_page, "off": (page - 1) * per_page}
    if condicion and condicion in _CONDICIONES_VALIDAS:
        where.append("condicion = :cond"); params["cond"] = condicion
    if aporta == "si":
        where.append("COALESCE(aporta_jdccpp, TRUE) = TRUE")
    elif aporta == "no":
        where.append("COALESCE(aporta_jdccpp, TRUE) = FALSE")
    if q and q.strip():
        where.append("(codigo_matricula ILIKE :q OR dni ILIKE :q OR apellidos_nombres ILIKE :q)")
        params["q"] = f"%{q.strip()}%"
    w = " AND ".join(where)
    total = db.execute(text(f"SELECT COUNT(*) AS c FROM colegiados WHERE {w}"), params).fetchone().c
    filas = db.execute(text(f"""
        SELECT id, codigo_matricula, apellidos_nombres, dni, condicion,
               COALESCE(aporta_jdccpp, TRUE) AS aporta_jdccpp,
               aporta_jdccpp_motivo AS motivo
        FROM colegiados WHERE {w}
        ORDER BY apellidos_nombres LIMIT :lim OFFSET :off
    """), params).fetchall()
    return JSONResponse({
        "total": total, "page": page, "pages": (total + per_page - 1) // per_page,
        "rows": [{
            "id": r.id, "codigo_matricula": r.codigo_matricula,
            "apellidos_nombres": r.apellidos_nombres, "dni": r.dni,
            "condicion": r.condicion, "aporta_jdccpp": bool(r.aporta_jdccpp),
            "motivo": r.motivo,
        } for r in filas],
    })


@router.put("/padron/{colegiado_id}/condicion")
async def padron_set_condicion(
    colegiado_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    """Cambia la condición del colegiado y AUTO-GESTIONA aporta_jdccpp.

    ── Regla de negocio CCPL (auto-gestión del flag aporta_jdccpp) ──
    'Past Decano' es un estado VIRTUAL (no es una condición real en BD): se guarda
    como condicion='habil' + aporta_jdccpp=FALSE + motivo='Past Decano'.
    aporta_jdccpp se deriva de la condición, EXCEPTO cuando el motivo es 'Past Decano'
    (que se preserva mientras la condición efectiva sea hábil):
      - 'past_decano'      → habil,            aporta=FALSE, motivo='Past Decano'
      - 'habil'            → habil,            aporta=TRUE,  motivo=NULL
                             (si el motivo previo era 'Past Decano', se PRESERVA:
                              queda habil, aporta=FALSE, motivo='Past Decano')
      - 'inhabil'          → inhabil,          aporta=FALSE, motivo='Inhábil'
      - 'vitalicio'        → vitalicio,        aporta=FALSE, motivo='Vitalicio'
      - 'candidato_retiro' → candidato_retiro, aporta=FALSE, motivo='Candidato a retiro'
    """
    t = (payload.get("condicion") or "").strip()
    if t not in _CONDICIONES_VALIDAS and t != "past_decano":
        raise HTTPException(400, "Condición no válida")

    actual = db.execute(text("""
        SELECT aporta_jdccpp_motivo AS motivo FROM colegiados
        WHERE id = :cid AND organization_id = :org
    """), {"cid": colegiado_id, "org": ORG_CCPL}).fetchone()
    if not actual:
        raise HTTPException(404, "Colegiado no encontrado")
    cur_motivo = actual.motivo

    if t == "past_decano":
        condicion, aporta, motivo = "habil", False, "Past Decano"
    elif t == "habil":
        if cur_motivo == "Past Decano":     # preservar Past Decano
            condicion, aporta, motivo = "habil", False, "Past Decano"
        else:
            condicion, aporta, motivo = "habil", True, None
    elif t == "inhabil":
        condicion, aporta, motivo = "inhabil", False, "Inhábil"
    elif t == "vitalicio":
        condicion, aporta, motivo = "vitalicio", False, "Vitalicio"
    else:  # candidato_retiro
        condicion, aporta, motivo = "candidato_retiro", False, "Candidato a retiro"

    db.execute(text("""
        UPDATE colegiados SET
            condicion = :c, fecha_actualizacion_condicion = NOW(),
            aporta_jdccpp = :a, aporta_jdccpp_motivo = :m,
            aporta_jdccpp_actualizado_por = :uid, aporta_jdccpp_actualizado_at = NOW()
        WHERE id = :cid AND organization_id = :org
    """), {"c": condicion, "a": aporta, "m": motivo, "uid": current_member.user_id,
           "cid": colegiado_id, "org": ORG_CCPL})
    db.commit()
    efectiva = "past_decano" if (condicion == "habil" and not aporta and motivo == "Past Decano") else condicion
    return JSONResponse({"ok": True, "condicion_efectiva": efectiva, "aporta_jdccpp": aporta})


@router.put("/padron/{colegiado_id}/aporta-jdccpp")
async def padron_set_aporta(
    colegiado_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    aporta = bool(payload.get("aporta"))
    motivo = (payload.get("motivo") or None)
    r = db.execute(text("""
        UPDATE colegiados SET
            aporta_jdccpp = :a, aporta_jdccpp_motivo = :m,
            aporta_jdccpp_actualizado_por = :uid, aporta_jdccpp_actualizado_at = NOW()
        WHERE id = :cid AND organization_id = :org
        RETURNING id
    """), {"a": aporta, "m": motivo, "uid": current_member.user_id,
           "cid": colegiado_id, "org": ORG_CCPL}).fetchone()
    if not r:
        raise HTTPException(404, "Colegiado no encontrado")
    db.commit()
    return JSONResponse({"ok": True})


# ════════════════════════════════════════════════════════════════
# PIEZA E — APROBACIÓN / REVOCACIÓN DE PERIODO
# ════════════════════════════════════════════════════════════════
@router.post("/periodo/{periodo_id}/aprobar")
async def aportes_aprobar(
    periodo_id: int,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo = db.execute(text("""
        SELECT id FROM aporte_periodos WHERE id = :pid AND organizacion_id = :org
    """), {"pid": periodo_id, "org": ORG_CCPL}).fetchone()
    if not periodo:
        raise HTTPException(404, "Periodo no encontrado")

    user = db.query(User).filter(User.id == current_member.user_id).first()
    dni = user.public_id if user else None
    nombre = user.name if user else None
    matricula = None
    if dni:
        col = db.execute(text("""
            SELECT codigo_matricula FROM colegiados
            WHERE dni = :dni AND organization_id = :org LIMIT 1
        """), {"dni": dni, "org": ORG_CCPL}).fetchone()
        matricula = col.codigo_matricula if col else None

    db.execute(text("""
        UPDATE aporte_periodos SET
            aprobado = TRUE, aprobado_por_user_id = :uid, aprobado_at = NOW(),
            revocado_por_user_id = NULL, revocado_at = NULL,
            aprobado_admin_dni = :dni, aprobado_admin_nombre = :nom, aprobado_admin_matricula = :mat,
            updated_at = NOW()
        WHERE id = :pid
    """), {"uid": current_member.user_id, "dni": dni, "nom": nombre, "mat": matricula, "pid": periodo_id})
    db.commit()
    return RedirectResponse(url=f"/admin/aportes-junta/periodo/{periodo_id}", status_code=303)


@router.post("/periodo/{periodo_id}/revocar")
async def aportes_revocar(
    periodo_id: int,
    db: Session = Depends(get_db),
    current_member: Member = Depends(require_aportes),
):
    periodo = db.execute(text("""
        SELECT id FROM aporte_periodos WHERE id = :pid AND organizacion_id = :org
    """), {"pid": periodo_id, "org": ORG_CCPL}).fetchone()
    if not periodo:
        raise HTTPException(404, "Periodo no encontrado")
    db.execute(text("""
        UPDATE aporte_periodos SET
            aprobado = FALSE, revocado_por_user_id = :uid, revocado_at = NOW(), updated_at = NOW()
        WHERE id = :pid
    """), {"uid": current_member.user_id, "pid": periodo_id})
    db.commit()
    return RedirectResponse(url=f"/admin/aportes-junta/periodo/{periodo_id}", status_code=303)

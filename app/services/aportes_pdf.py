"""
app/services/aportes_pdf.py
Generadores compartidos de PDF/Excel del reporte de aportes a la JDCCPP.

Usados por:
- Admin (app/routers/aportes_junta.py): PDF/Excel del periodo (borrador o firmado).
- Representante JDCCPP (app/routers/junta.py): descarga oficial (solo aprobados).

Pieza H: si el periodo está APROBADO, el PDF incluye el bloque de firma del
Administrador (nombre/DNI/matrícula/fecha de aprobación tomados del periodo).
Si NO está aprobado, se estampa marca de agua "PREVIEW — PENDIENTE DE APROBACIÓN".
"""

import io
from datetime import datetime, timezone, timedelta

from sqlalchemy import text
from sqlalchemy.orm import Session

TZ_PERU = timezone(timedelta(hours=-5))
MESES_ES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _fetch(db: Session, periodo_id: int, org_id: int = 1):
    periodo = db.execute(text("""
        SELECT ap.*, dep.numero_voucher, dep.fecha_deposito, dep.monto AS deposito_monto,
               dep.banco_emisor
        FROM aporte_periodos ap
        LEFT JOIN aporte_deposito dep ON dep.aporte_periodo_id = ap.id
        WHERE ap.id = :pid AND ap.organizacion_id = :org
    """), {"pid": periodo_id, "org": org_id}).fetchone()
    if not periodo:
        return None, None
    nuevos = db.execute(text("""
        SELECT codigo_matricula, apellidos_nombres, dni, fecha_pago_der_col, monto_aporte
        FROM aporte_detalle_nuevos WHERE aporte_periodo_id = :pid
        ORDER BY codigo_matricula NULLS LAST, apellidos_nombres
    """), {"pid": periodo_id}).fetchall()
    return periodo, nuevos


def generar_pdf(db: Session, periodo_id: int, show_footer: bool = False, org_id: int = 1):
    periodo, nuevos = _fetch(db, periodo_id, org_id)
    if not periodo:
        return None

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    aprobado = bool(getattr(periodo, "aprobado", False))
    periodo_label = f"{MESES_ES[periodo.mes]} {periodo.anio}"

    def _watermark(canvas, doc):
        if aprobado:
            return
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 40)
        canvas.setFillColor(colors.Color(0.85, 0.15, 0.15, alpha=0.16))
        canvas.translate(A4[0] / 2, A4[1] / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "PREVIEW — PENDIENTE DE APROBACIÓN")
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=14)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=11, alignment=1)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.grey, alignment=1)

    el = []
    el.append(Paragraph("COLEGIO DE CONTADORES PÚBLICOS DE LORETO", h))
    el.append(Paragraph(f"Resumen de depósito mensual a JDCCPP — {periodo_label.upper()}", sub))
    el.append(Spacer(1, 10 * mm))

    resumen = [
        ["TOTAL CUOTAS ORDINARIAS MIEMBROS HÁBILES", f"S/ {float(periodo.monto_habiles or 0):,.2f}"],
        ["TOTAL NUEVOS COLEGIADOS", f"S/ {float(periodo.monto_nuevos or 0):,.2f}"],
        [f"TOTAL A DEPOSITAR A JDCCPP — {periodo_label.upper()}", f"S/ {float(periodo.monto_total or 0):,.2f}"],
    ]
    t = Table(resumen, colWidths=[120 * mm, 50 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, -1), (-1, -1), 1, colors.black),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    el.append(t)
    el.append(Spacer(1, 8 * mm))

    el.append(Paragraph("DETALLE DE NUEVOS COLEGIADOS", styles["Heading4"]))
    data = [["N°", "Matrícula", "Apellidos y Nombres", "DNI", "F. Pago", "Monto"]]
    for i, n in enumerate(nuevos, 1):
        data.append([
            str(i), n.codigo_matricula or "—", n.apellidos_nombres, n.dni or "—",
            n.fecha_pago_der_col.strftime("%d/%m/%Y") if n.fecha_pago_der_col else "—",
            f"S/ {float(n.monto_aporte or 0):,.2f}",
        ])
    td = Table(data, colWidths=[10 * mm, 22 * mm, 78 * mm, 24 * mm, 20 * mm, 24 * mm], repeatRows=1)
    td.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ALIGN", (0, 0), (0, -1), "CENTER"), ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    el.append(td)
    el.append(Spacer(1, 10 * mm))

    if periodo.numero_voucher:
        el.append(Paragraph(
            f"Depósito: S/ {float(periodo.deposito_monto or 0):,.2f} · Voucher {periodo.numero_voucher}"
            f" · {periodo.banco_emisor or ''}"
            + (f" · {periodo.fecha_deposito.strftime('%d/%m/%Y')}" if periodo.fecha_deposito else ""),
            styles["Normal"]))
        el.append(Spacer(1, 8 * mm))

    # ── Pieza H: bloque de firma (solo si aprobado) ──
    hoy = datetime.now(TZ_PERU).strftime("%d de %B de %Y")
    if aprobado:
        aprob_at = getattr(periodo, "aprobado_at", None)
        aprob_str = aprob_at.astimezone(TZ_PERU).strftime("%d/%m/%Y %H:%M") if aprob_at else "—"
        nombre_admin = getattr(periodo, "aprobado_admin_nombre", None)
        firma = ParagraphStyle("firma", parent=styles["Normal"], fontSize=10, leading=15)
        el.append(Spacer(1, 6 * mm))
        el.append(Paragraph(f"Iquitos, {hoy}", styles["Normal"]))
        el.append(Spacer(1, 8 * mm))
        el.append(Paragraph("Aprobado y publicado por:", styles["Normal"]))
        el.append(Spacer(1, 3 * mm))
        if nombre_admin:
            el.append(Paragraph(
                f"<b>{nombre_admin}</b><br/>"
                f"Administrador del CCPL<br/>"
                f"DNI: {getattr(periodo,'aprobado_admin_dni',None) or '—'}<br/>"
                f"Matrícula: {getattr(periodo,'aprobado_admin_matricula',None) or '—'}<br/>"
                f"Fecha de aprobación: {aprob_str}", firma))
        else:
            # Aprobación retroactiva por el sistema (ej. Mayo 2026 aprobado vía SQL,
            # sin firmante). No hay DNI/matrícula que mostrar.
            el.append(Paragraph(
                f"<b>Regularización retroactiva del sistema</b><br/>"
                f"Fecha de aprobación: {aprob_str}", firma))
    else:
        el.append(Paragraph(
            "<b>DOCUMENTO PREVIEW — PENDIENTE DE APROBACIÓN</b>",
            ParagraphStyle("pv", parent=styles["Normal"], textColor=colors.HexColor("#b91c1c"))))

    if show_footer:
        el.append(Spacer(1, 12 * mm))
        el.append(Paragraph(
            "Sistema desarrollado por Perú Sistemas Pro · perusistemas.pro · WhatsApp +51 967 317 946",
            small))

    doc.build(el, onFirstPage=_watermark, onLaterPages=_watermark)
    return buf.getvalue()


def generar_excel(db: Session, periodo_id: int, show_footer: bool = False, org_id: int = 1):
    periodo, nuevos = _fetch(db, periodo_id, org_id)
    if not periodo:
        return None

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    periodo_label = f"{MESES_ES[periodo.mes]} {periodo.anio}"
    azul = PatternFill("solid", fgColor="1E3A5F")
    blanco_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"
    ws1["A1"] = f"Aporte a JDCCPP — {periodo_label}"; ws1["A1"].font = Font(bold=True, size=14)
    rows = [
        ("", ""),
        ("Cuotas ordinarias hábiles", float(periodo.monto_habiles or 0)),
        (f"   ({periodo.cantidad_habiles or 0} hábiles)", ""),
        ("Nuevos colegiados", float(periodo.monto_nuevos or 0)),
        (f"   ({periodo.cantidad_nuevos or 0} nuevos)", ""),
        ("TOTAL A DEPOSITAR", float(periodo.monto_total or 0)),
        ("", ""),
        ("Estado", periodo.estado),
        ("Aprobado", "Sí" if getattr(periodo, "aprobado", False) else "No"),
        ("Voucher", periodo.numero_voucher or "—"),
        ("Depósito S/", float(periodo.deposito_monto or 0)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws1[f"A{i}"] = k; ws1[f"B{i}"] = v
    ws1["A8"].font = bold; ws1["B8"].font = bold
    ws1.column_dimensions["A"].width = 34; ws1.column_dimensions["B"].width = 18

    ws2 = wb.create_sheet("Detalle Nuevos")
    headers = ["N°", "Matrícula", "Apellidos y Nombres", "DNI", "Fecha Pago", "Monto Aporte"]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c); cell.fill = azul; cell.font = blanco_bold
        cell.alignment = Alignment(horizontal="center")
    for i, n in enumerate(nuevos, 1):
        ws2.append([
            i, n.codigo_matricula or "—", n.apellidos_nombres, n.dni or "—",
            n.fecha_pago_der_col.strftime("%d/%m/%Y") if n.fecha_pago_der_col else "—",
            float(n.monto_aporte or 0),
        ])
    for col, w in zip("ABCDEF", [6, 14, 44, 14, 14, 14]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Metadata")
    meta = [
        ("Periodo", periodo_label),
        ("Generado", datetime.now(TZ_PERU).strftime("%d/%m/%Y %H:%M")),
        ("Total nuevos", periodo.cantidad_nuevos or 0),
        ("Total hábiles", periodo.cantidad_habiles or 0),
        ("UIT aplicada", float(periodo.uit_aplicada or 0)),
        ("Total a depositar", float(periodo.monto_total or 0)),
        ("Aprobado", "Sí" if getattr(periodo, "aprobado", False) else "No"),
        ("Aprobado por", getattr(periodo, "aprobado_admin_nombre", None) or "—"),
        ("DNI firmante", getattr(periodo, "aprobado_admin_dni", None) or "—"),
        ("Voucher", periodo.numero_voucher or "—"),
        ("Marco normativo", "Acuerdo institucional JDCCPP"),
    ]
    for k, v in meta:
        ws3.append([k, v])
    if show_footer:
        ws3.append(["", ""])
        ws3.append(["Sistema", "Perú Sistemas Pro · perusistemas.pro · WhatsApp +51 967 317 946"])
    ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 48

    out = io.BytesIO(); wb.save(out)
    return out.getvalue()

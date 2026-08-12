"""
app/services/export_comprobantes_dia.py
Export a Excel de los comprobantes del día (sesión de caja actual), UNA FILA POR
CONCEPTO. Primera pieza del módulo de Consultas de Caja.

13 columnas (orden fijo):
  FECHA | MATRICULA | APELLIDOS Y NOMBRES | COD_CDP | CDP | NUM_CPD |
  CONCEPTO INGRESO | CONCEPTO | OBSERVACION | COBRANZA | PERIODO |
  FORMA PAGO | IMPORTE

Fuente de comprobantes: tabla `comprobantes` (modelo Comprobante), la misma que
usa el arqueo. El desglose por concepto se reconstruye por DEUDA (payment_debts /
[DEBT_IDS:] en notes → debts → conceptos_cobro) como vía universal para caja y
OpenPay; CONCEPTOS_B64 se usa como complemento para conceptos sin deuda
(constancias, eventos con genera_deuda=False). El notes JSON de OpenPay NO trae
importe por concepto, por eso NO se usa como fuente del desglose.

Solo lectura de BD; no muta nada.
"""

import io
import re
import json
import base64
from datetime import datetime, timezone, timedelta

from sqlalchemy import text, bindparam
from sqlalchemy.orm import Session

TZ_PERU = timezone(timedelta(hours=-5))

CDP_MAP = {
    "01": "FACTURA",
    "03": "BOLETA DE VENTA",
    "07": "NOTA DE CREDITO",
    "08": "NOTA DE DEBITO",
}

COBRANZA_WEB = "WEB"   # pagos OpenPay online (sin sesión de caja). Cambiar a "OPENPAY" si se prefiere.

_PLACEHOLDERS_REF = {"", "pendiente", "por validar en voucher", "por validar en foto"}
_RE_DEBT_IDS = re.compile(r"\[DEBT_IDS:([0-9,\s]+)\]")
_RE_B64 = re.compile(r"\[CONCEPTOS_B64:([A-Za-z0-9+/=]+)\]")


def _norm_forma_pago(pm: str) -> str:
    s = (pm or "").strip().lower()
    if not s:
        return ""
    if "efectivo" in s:
        return "Efectivo"
    if "yape" in s:
        return "Yape"
    if "plin" in s:
        return "Plin"
    if "transfer" in s:
        return "Transferencia"
    if "openpay" in s:
        return "OpenPay"
    if any(k in s for k in ("tarjeta", "card", "visa", "master", "pos")):
        return "Tarjeta"
    return (pm or "").strip().title()


def _ref_pago(operation_code: str, openpay_tx: str) -> str:
    oc = (operation_code or "").strip()
    if oc and oc.lower() not in _PLACEHOLDERS_REF:
        return oc
    return (openpay_tx or "").strip()


def _parse_b64(notes: str):
    if not notes:
        return []
    m = _RE_B64.search(notes)
    if not m:
        return []
    try:
        decoded = base64.b64decode(m.group(1)).decode("utf-8")
        data = json.loads(decoded)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _conceptos_de_pago(db: Session, org_id: int, pay_id, notes: str):
    """Lista de {ci, concepto, importe(+)} por concepto. Reconstruye por deuda y
    complementa con CONCEPTOS_B64 para conceptos sin deuda."""
    rows = []
    codigos_con_deuda = set()

    debt_ids = set()
    if pay_id:
        for r in db.execute(text("SELECT debt_id FROM payment_debts WHERE payment_id = :p"),
                            {"p": pay_id}).fetchall():
            if r[0] is not None:
                debt_ids.add(int(r[0]))
    m = _RE_DEBT_IDS.search(notes or "")
    if m:
        for x in m.group(1).split(","):
            x = x.strip()
            if x.isdigit():
                debt_ids.add(int(x))

    if debt_ids:
        q = text("""
            SELECT d.concept, d.amount, cc.codigo AS cod, cc.nombre AS cc_nombre,
                   cc.nombre_corto AS cc_corto
            FROM debts d
            LEFT JOIN conceptos_cobro cc ON cc.id = d.concepto_cobro_id
            WHERE d.id IN :ids
        """).bindparams(bindparam("ids", expanding=True))
        for d in db.execute(q, {"ids": list(debt_ids)}).fetchall():
            ci = d.cc_corto or d.cc_nombre or (d.concept or "")
            if d.cod:
                codigos_con_deuda.add(d.cod)
            rows.append({
                "ci": ci,
                "concepto": d.concept or d.cc_nombre or "",
                "importe": float(d.amount or 0),
            })

    # Complemento: conceptos del B64 que no tienen deuda (constancias, etc.)
    for c in _parse_b64(notes):
        cod = (c.get("codigo") or "").strip()
        if cod and cod in codigos_con_deuda:
            continue
        cc = None
        if cod:
            cc = db.execute(text("""
                SELECT nombre, nombre_corto FROM conceptos_cobro
                WHERE organization_id = :o AND codigo = :c LIMIT 1
            """), {"o": org_id, "c": cod}).fetchone()
        ci = (cc.nombre_corto if cc and cc.nombre_corto else c.get("nombre")) or ""
        concepto = (cc.nombre if cc and cc.nombre else None) or c.get("nombre") or ""
        rows.append({
            "ci": ci,
            "concepto": concepto,
            "importe": float(c.get("monto_total") or 0),
        })

    return rows


def generar_export_comprobantes(db: Session, org_id: int, desde) -> bytes:
    """Genera el .xlsx de comprobantes emitidos desde `desde` (inclusive).
    Incluye 01/03/07/08 con estado accepted/anulado. NC (07) → IMPORTE negativo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    comps = db.execute(text("""
        SELECT c.id AS comp_id, c.tipo, c.serie, c.numero, c.total, c.observaciones,
               c.created_at AS comp_created, c.payment_id,
               p.id AS pay_id, p.created_at AS pay_created, p.payment_method,
               p.operation_code, p.openpay_transaction_id, p.notes,
               p.colegiado_id,
               col.codigo_matricula, col.apellidos_nombres
        FROM comprobantes c
        LEFT JOIN payments p ON p.id = c.payment_id
        LEFT JOIN colegiados col ON col.id = p.colegiado_id
        WHERE c.organization_id = :org
          AND c.created_at >= :desde
          AND c.tipo IN ('01', '03', '07', '08')
          AND c.status IN ('accepted', 'anulado')
        ORDER BY c.created_at, c.id
    """), {"org": org_id, "desde": desde}).fetchall()

    headers = [
        "FECHA", "MATRICULA", "APELLIDOS Y NOMBRES", "COD_CDP", "CDP", "NUM_CPD",
        "CONCEPTO INGRESO", "CONCEPTO", "OBSERVACION", "COBRANZA", "PERIODO",
        "FORMA PAGO", "IMPORTE",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes del día"
    ws.append(headers)
    azul = PatternFill("solid", fgColor="1E3A5F")
    blanco_bold = Font(color="FFFFFF", bold=True)
    for col_i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = azul
        cell.font = blanco_bold
        cell.alignment = Alignment(horizontal="center")

    for c in comps:
        forma_pago = _norm_forma_pago(c.payment_method)
        ref = _ref_pago(c.operation_code, c.openpay_transaction_id)
        observacion = f"{forma_pago.upper()} REF:{ref}" if ref else ""
        notes = c.notes or ""
        cobranza = "VENTANILLA" if notes.lstrip().startswith("[CAJA]") else COBRANZA_WEB
        cod_cdp = c.tipo or ""
        cdp = CDP_MAP.get(c.tipo, c.tipo or "")
        num_cpd = f"{c.serie or ''}-{int(c.numero or 0):08d}"
        matricula = c.codigo_matricula or ""
        nombres = c.apellidos_nombres or ""

        fecha_dt = c.pay_created or c.comp_created
        if fecha_dt is not None:
            fecha_peru = fecha_dt.astimezone(TZ_PERU) if fecha_dt.tzinfo else fecha_dt
            fecha_str = fecha_peru.strftime("%d/%m/%Y")
            periodo_str = fecha_peru.strftime("%Y%m%d")
        else:
            fecha_str = periodo_str = ""

        signo = -1 if c.tipo == "07" else 1
        conceptos = _conceptos_de_pago(db, org_id, c.pay_id, notes)
        if not conceptos:
            conceptos = [{
                "ci": "",
                "concepto": c.observaciones or cdp,
                "importe": float(c.total or 0),
            }]

        for cc in conceptos:
            ws.append([
                fecha_str, matricula, nombres, cod_cdp, cdp, num_cpd,
                cc["ci"], cc["concepto"], observacion, cobranza, periodo_str,
                forma_pago, round(signo * float(cc["importe"] or 0), 2),
            ])

    anchos = [12, 12, 40, 8, 18, 18, 22, 40, 26, 12, 12, 14, 12]
    for i, w in enumerate(anchos):
        ws.column_dimensions[chr(64 + i + 1) if i < 26 else "A"].width = w
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
